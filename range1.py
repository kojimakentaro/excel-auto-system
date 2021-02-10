#!/usr/bin/env python
# coding: utf-8

# In[1]:


import openpyxl


# In[3]:


wb = openpyxl.Workbook()
wb.active
wb.save("test.xlsx")


# In[4]:


wb.sheetnames


# In[6]:


ws =wb["Sheet"]


# In[7]:


ws.title


# In[8]:


cell1 =ws["A1"]


# In[9]:


cell2 =ws.cell(row=1,column=1)


# In[10]:


cell3 =ws.cell(1,1)


# In[12]:


cell1==cell2==cell3


# In[16]:


range1=ws["A1":"B3"]

range2=ws["A1:B3"]


# In[17]:


range1==range2


# In[18]:


range1


# In[19]:


range1[0]


# In[20]:


range1[1]


# In[21]:


for row in range1:
    for cell in row:
        print(cell)


# In[23]:


for row in range1:
    format_l =[ ]
    for cell in row:
        format_l.append(cell.coordinate)
        print(format_l)


# In[26]:


def show_formatted_cell(range_):
    for row in range_:
        format_l =[ ]
        for cell in row:
            format_l.append(cell.coordinate)
        print(format_l)


# In[27]:


show_formatted_cell(range1)


# In[29]:


print(wb.sheetnames)
ws =wb["Sheet"]


# In[30]:


range1 =ws["A1:C6"]


# In[31]:


def show_formatted_cell(range1):
    for row in range1:
        format_l =[ ]
        for cell in row:
            format_l.append(cell.coordinate)
        print(format_l)


# In[32]:


show_formatted_cell(range1)


# In[38]:


for row in ws.iter_rows(min_row=2):
    format_l=[]
    for cell in row:
        format_l.append(cell.coordinate)
    print(format_l)


# In[39]:


for row in ws.iter_rows(min_row=2,min_col=2):
    format_l=[]
    for cell in row:
        format_l.append(cell.coordinate)
    print(format_l)


# In[45]:


def show_formated_cell(ws, min_row=1,min_col=1):
    for row in ws.iter_rows(min_row,min_col):
        format_l=[]
        for cell in row:
            format_l.append(cell.coordinate)
            print(format_l)


# In[48]:


show_formated_cell(ws)

show_formated_cell(ws,2,2)


# In[49]:


row1 =ws[1]


# In[50]:


row1


# In[51]:


row1[0]


# In[53]:


cell1 =ws.cell(row=1,column=1)


# In[61]:


print(cell1.coordinate)
print(cell1.row)
print(cell1.column)
print(cell1.column_letter)


# In[63]:


ws = wb["Sheet"]

cell1 =ws.cell(row=1,column=1)
print(cell1)


# In[64]:


cell1.value=1
print(cell1.value)


# In[65]:


wb.save("test.xlsx")


# In[67]:


cell2 =ws.cell(row=2,column =1)
cell2.value=2
print("セルの場所",cell2.coordinate)
print("セルの値",cell2.value)


# In[68]:


wb.save("test.xlsx")


# In[71]:


cell3 =ws.cell(row=3,column=1,value=10)
print(cell3.coordinate)
print(cell3.value)
wb.save("test.xlsx")


# In[78]:


cell4 = ws.cell(row =1,column=2)
print(cell4.coordinate)
cell4.value = "SUM(A1:A3)"
wb.save("test.xlsx")


# In[79]:


for row in ws.values:
    print(row)


# In[81]:


cell5 =ws.cell(row =2,column=2,value=1.2345)
print(cell5.coordinate)
print(cell5.value)


# In[82]:


from datetime import date


# In[83]:


cell6 =ws.cell(row=3,column=2,value=date(2020,10,10))
print(cell6.coordinate)
print(cell6.value)


# In[84]:


wb.save("test.xlsx")


# In[89]:


ws["B2"].number_format="0.00"
ws["B3"].number_format="yyyy年mm月dd日"
print(ws["B2"].number_format)
print(ws["B3"].number_format)


# In[ ]:




